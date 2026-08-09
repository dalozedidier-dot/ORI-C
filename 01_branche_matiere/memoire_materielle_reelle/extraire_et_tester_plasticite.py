#!/usr/bin/env python3
"""FABEST et aciers moyen-Mn : extraction et test de dose d'histoire.

FABEST : corrélation cycle-amplitude, une valeur par éprouvette,
24 éprouvettes en trois groupes de chargement.
Moyen-Mn : corrélation maintien-dureté à température constante.

    python extraire_et_tester_plasticite.py
"""
from __future__ import annotations

import csv
import json
import warnings
from collections import defaultdict
from pathlib import Path

import numpy as np

warnings.filterwarnings("ignore")

ICI = Path(__file__).resolve().parent
SOURCES = ICI / "SOURCES.json"
DERIVE = ICI / "derive"
ALPHA = 0.05
TIRAGES = 10000
GRAINE = 20260809


def spearman(x, y) -> float:
    x, y = np.asarray(x, float), np.asarray(y, float)
    if x.size < 3:
        return float("nan")
    def rangs(v):
        o = np.argsort(v, kind="mergesort")
        r = np.empty(len(v), dtype=float)
        r[o] = np.arange(len(v), dtype=float)
        return r
    rx, ry = rangs(x), rangs(y)
    if rx.std() == 0 or ry.std() == 0:
        return float("nan")
    return float(np.corrcoef(rx, ry)[0, 1])


def p_sign_flip(valeurs: np.ndarray, aleatoire) -> tuple[float, str]:
    n = valeurs.size
    if n == 0:
        return 1.0, "aucune valeur"
    observe = abs(float(valeurs.mean()))
    if n <= 20:
        compte = sum(
            1 for masque in range(1 << n)
            if abs(float((valeurs * np.array(
                [1.0 if masque >> i & 1 else -1.0 for i in range(n)])).mean())) >= observe)
        return compte / (1 << n), f"sign-flip exact, {1 << n} attributions"
    compte = sum(1 for _ in range(TIRAGES)
                 if abs(float((valeurs * aleatoire.choice((-1.0, 1.0), size=n)).mean())) >= observe)
    return (1 + compte) / (1 + TIRAGES), f"sign-flip Monte-Carlo, {TIRAGES} tirages"


# ------------------------------------------------------------------ FABEST

def extraire_fabest(racine: Path) -> list[dict]:
    """Une ligne par éprouvette : corrélation cycle / amplitude de contrainte."""
    base = racine / "fabest_lcf" / "exploitable"
    eprouvettes = []
    for fichier in sorted(base.rglob("amp_mean_values.csv")):
        # Le correctif de juillet 2026 duplique des essais octet pour octet.
        if "update" in str(fichier).lower():
            continue
        parties = fichier.parts
        cas = next((p for p in parties if p.startswith("Case_")), None)
        essai = next((p for p in parties if p.startswith("Exp_")), None)
        if not cas or not essai:
            continue
        try:
            with fichier.open(encoding="utf-8", errors="replace", newline="") as flux:
                lignes = list(csv.reader(flux))
        except OSError:
            continue
        if len(lignes) < 4:
            continue
        entete = [c.strip().lower() for c in lignes[0]]
        colonne_cycle = next((i for i, c in enumerate(entete) if "cycle" in c), 0)
        colonne_amp = next((i for i, c in enumerate(entete)
                            if "amp" in c and "stress" in c), None)
        if colonne_amp is None:
            colonne_amp = next((i for i, c in enumerate(entete) if "amp" in c), 1)
        cycles, amplitudes = [], []
        for ligne in lignes[1:]:
            try:
                cycles.append(float(ligne[colonne_cycle]))
                amplitudes.append(float(ligne[colonne_amp]))
            except (ValueError, IndexError):
                continue
        if len(cycles) < 5:
            continue
        eprouvettes.append({
            "source": "fabest_lcf", "groupe": cas, "eprouvette": essai,
            "n_cycles": len(cycles),
            "amplitude_debut": amplitudes[0], "amplitude_fin": amplitudes[-1],
            "rho_cycle_amplitude": spearman(cycles, amplitudes),
            "data_kind": "mesure_experimentale",
        })
    return eprouvettes


# ------------------------------------------------------------ moyen-Mn

def extraire_medium_mn(racine: Path) -> list[dict]:
    """Une ligne par éprouvette : histoire thermique et dureté moyenne."""
    from openpyxl import load_workbook
    chemins = list((racine / "medium_mn_a" / "exploitable").rglob("twardosc.xlsx"))
    if not chemins:
        return []
    classeur = load_workbook(chemins[0], read_only=True, data_only=True)
    feuille = classeur[classeur.sheetnames[0]]
    lignes = list(feuille.iter_rows(values_only=True))
    classeur.close()
    if len(lignes) < 6:
        return []
    temperatures, durees, identifiants = lignes[2], lignes[3], lignes[4]
    eprouvettes = []
    for colonne in range(1, len(temperatures)):
        identifiant = identifiants[colonne] if colonne < len(identifiants) else None
        temperature = temperatures[colonne]
        duree = durees[colonne] if colonne < len(durees) else None
        if not identifiant or not isinstance(temperature, (int, float)):
            continue
        if not isinstance(duree, (int, float)):
            continue
        mesures = [l[colonne] for l in lignes[5:]
                   if colonne < len(l) and isinstance(l[colonne], (int, float))]
        if len(mesures) < 3:
            continue
        eprouvettes.append({
            "source": "medium_mn_a", "eprouvette": str(identifiant),
            "temperature_C": float(temperature), "maintien_s": float(duree),
            "n_indentations": len(mesures),
            "durete_moyenne_HV": float(np.mean(mesures)),
            "durete_ecart_type": float(np.std(mesures, ddof=1)),
            "data_kind": "mesure_experimentale",
        })
    return eprouvettes


def main() -> int:
    config = json.loads(SOURCES.read_text(encoding="utf-8"))
    racine = (ICI / config["racine_locale"]).resolve()
    attendues = [racine / "fabest_lcf" / "exploitable",
                 racine / "medium_mn_a" / "exploitable"]
    absentes = [d for d in attendues if not d.is_dir()]
    if absentes:
        for dossier in absentes:
            print(f"source absente : {dossier}")
        print("Résultats commités laissés intacts.")
        return 1
    DERIVE.mkdir(exist_ok=True)
    aleatoire = np.random.default_rng(GRAINE)
    rapports = {}

    # ---------------------------------------------------------------- FABEST
    fabest = extraire_fabest(racine)
    if fabest:
        with (DERIVE / "fabest_par_eprouvette.csv").open(
                "w", encoding="utf-8", newline="") as flux:
            redacteur = csv.DictWriter(flux, fieldnames=list(fabest[0]), lineterminator="\n")
            redacteur.writeheader()
            redacteur.writerows(fabest)
        groupes = defaultdict(list)
        for e in fabest:
            groupes[e["groupe"]].append(e["rho_cycle_amplitude"])
        rhos = np.array([e["rho_cycle_amplitude"] for e in fabest
                         if np.isfinite(e["rho_cycle_amplitude"])])
        p, methode = p_sign_flip(rhos, aleatoire)
        rapports["fabest_lcf"] = {
            "famille": "plasticite",
            "plan": "groupes_independants",
            "eprouvettes": len(fabest),
            "groupes": {g: len(v) for g, v in sorted(groupes.items())},
            "rho_moyen_cycle_amplitude": float(rhos.mean()),
            "fraction_negative": float((rhos < 0).mean()),
            "p": p, "methode_du_temoin": methode,
            "verdict": ("soutient" if p <= ALPHA else "ne_soutient_pas"),
            "lecture": ("l'amplitude de contrainte évolue avec le nombre de cycles "
                        "accumulés : le cyclage inscrit. Une valeur par éprouvette, "
                        "jamais une par cycle."),
        }
        print(f"FABEST : {len(fabest)} éprouvettes, "
              f"rho moyen {rhos.mean():+.4f}, p = {p:.4f}, "
              f"{rapports['fabest_lcf']['verdict']}")
        for g, v in sorted(groupes.items()):
            print(f"    {g:<34}{len(v):>3} éprouvettes  rho moyen {np.mean(v):+.4f}")
    else:
        print("FABEST : aucune éprouvette extraite")

    # -------------------------------------------------------------- moyen-Mn
    mmn = extraire_medium_mn(racine)
    if mmn:
        with (DERIVE / "medium_mn_par_eprouvette.csv").open(
                "w", encoding="utf-8", newline="") as flux:
            redacteur = csv.DictWriter(flux, fieldnames=list(mmn[0]), lineterminator="\n")
            redacteur.writeheader()
            redacteur.writerows(mmn)
        par_temperature = defaultdict(list)
        for e in mmn:
            par_temperature[e["temperature_C"]].append(e)
        rhos = []
        detail = {}
        for temperature, groupe in sorted(par_temperature.items()):
            if len(groupe) < 3:
                continue
            rho = spearman([g["maintien_s"] for g in groupe],
                           [g["durete_moyenne_HV"] for g in groupe])
            if np.isfinite(rho):
                rhos.append(rho)
                detail[f"{temperature:.0f}C"] = {"eprouvettes": len(groupe), "rho": rho}
        rhos = np.array(rhos)
        p, methode = p_sign_flip(rhos, aleatoire) if rhos.size else (1.0, "aucune")
        rapports["medium_mn_a"] = {
            "famille": "transition_de_phase",
            "plan": "factoriel temperature x maintien",
            "eprouvettes": len(mmn),
            "temperatures": detail,
            "rho_moyen_maintien_durete": float(rhos.mean()) if rhos.size else float("nan"),
            "p": p, "methode_du_temoin": methode,
            "verdict": ("soutient" if p <= ALPHA else
                        "indetermine_par_atteignabilite" if rhos.size < 6 else
                        "ne_soutient_pas"),
            "lecture": ("la dureté suit-elle la durée de maintien, à température "
                        "constante. Les neuf indentations d'une éprouvette sont des "
                        "mesures répétées, pas des unités."),
            "atteignabilite": (f"{rhos.size} températures exploitables ; sign-flip "
                               f"exact, p minimal 2/2**{rhos.size} = "
                               f"{2.0 / 2 ** rhos.size:.4f}" if rhos.size else "aucune"),
        }
        print()
        print(f"moyen-Mn : {len(mmn)} éprouvettes, {len(detail)} températures")
        for t, d in detail.items():
            print(f"    {t:<8}{d['eprouvettes']:>3} éprouvettes  rho {d['rho']:+.4f}")
        print(f"    p = {p:.4f}  →  {rapports['medium_mn_a']['verdict']}")
    else:
        print("moyen-Mn : aucune éprouvette extraite")

    sortie = DERIVE / "RESULTATS_PLASTICITE_ET_PHASE.json"
    with sortie.open("w", encoding="utf-8", newline="") as flux:
        flux.write(json.dumps({
            "campagne": "WP-MAT-MEM-2026", "alpha": ALPHA, "graine": GRAINE,
            "estimateur_de_p": "(1 + k) / (1 + N)",
            "jeux": rapports,
        }, ensure_ascii=False, indent=2) + "\n")
    print()
    print(f"écrit : {sortie.relative_to(ICI.parents[1]).as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
