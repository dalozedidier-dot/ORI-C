#!/usr/bin/env python3
"""Aciers à outils revenus : la chaîne complète histoire, trace, réponse.

Six éprouvettes, deux matériaux, trois températures de revenu chacun. C'est le
seul jeu de la campagne où les trois maillons sont mesurés sur les mêmes
éprouvettes :

    histoire  température du second revenu
    trace     dureté HRC, mesurée avant les essais de réponse
    réponse   ténacité à la rupture, coefficient de frottement

Statistique : corrélation de rang, stratifiée par matériau. Témoin : permutation
des températures à l'intérieur de chaque matériau.

    python extraire_et_tester_carbures.py
"""
from __future__ import annotations

import csv
import json
import math
import re
import warnings
from pathlib import Path

import numpy as np

from statistiques_rangs import spearman

from donnees_campagne import base_de

warnings.filterwarnings("ignore")

ICI = Path(__file__).resolve().parent
SOURCES = ICI / "SOURCES.json"
DERIVE = ICI / "derive"

ALPHA = 0.05
GRAINE = 20260809
TIRAGES = 20000
REVENU = re.compile(r"(\d{3})\s*°C")


def feuille(chemin: Path) -> list[tuple]:
    from openpyxl import load_workbook
    classeur = load_workbook(chemin, read_only=True, data_only=True)
    lignes = list(classeur[classeur.sheetnames[0]].iter_rows(values_only=True))
    classeur.close()
    return lignes


def extraire(base: Path) -> list[dict]:
    traitements = {}
    for ligne in feuille(base / "Heat treatments.xlsx")[1:]:
        if not ligne or not ligne[0]:
            continue
        sequence = str(ligne[4] or "")
        # La séquence enchaîne deux revenus, « 500°C/2h + 510°C/2h » : c'est le
        # dernier qui fixe l'état final et sert de dose.
        temperatures = REVENU.findall(sequence)
        if temperatures:
            traitements[str(ligne[0]).strip()] = float(temperatures[-1])

    reponses = {}
    for ligne in feuille(base / "Fracture Toughnes.xlsx")[1:]:
        if not ligne or not ligne[0]:
            continue
        try:
            reponses[str(ligne[0]).strip()] = (float(ligne[1]), float(ligne[2]))
        except (TypeError, ValueError):
            continue

    frottements = {}
    for nom, cle in (("COF high load .xlsx", "haute"), ("COF low load .xlsx", "basse")):
        chemin = base / nom
        if not chemin.exists():
            continue
        for ligne in feuille(chemin):
            if not ligne or not ligne[0]:
                continue
            try:
                frottements.setdefault(str(ligne[0]).strip(), {})[cle] = float(ligne[1])
            except (TypeError, ValueError):
                continue

    eprouvettes = []
    for identifiant, revenu in sorted(traitements.items()):
        if identifiant not in reponses:
            continue
        tenacite, durete = reponses[identifiant]
        eprouvettes.append({
            "source": "carbures_matrice_acier",
            "eprouvette": identifiant,
            "materiau": identifiant.split("-")[0],
            "revenu_C": revenu,
            "durete_HRC": durete,
            "tenacite": tenacite,
            "cof_haute_charge": frottements.get(identifiant, {}).get("haute", float("nan")),
            "cof_basse_charge": frottements.get(identifiant, {}).get("basse", float("nan")),
            "data_kind": "mesure_experimentale",
        })
    return eprouvettes


def tester(eprouvettes: list[dict], x: str, y: str, aleatoire) -> dict:
    """Corrélation stratifiée par matériau, testée par permutation interne."""
    strates = {}
    for e in eprouvettes:
        strates.setdefault(e["materiau"], []).append((e[x], e[y]))
    strates = {m: p for m, p in strates.items()
               if len(p) >= 3 and all(np.isfinite(a) and np.isfinite(b) for a, b in p)}
    if len(strates) < 2:
        return {"testable": False, "motif": f"{len(strates)} strate(s), minimum 2"}

    def statistique(s):
        rhos = [spearman([a for a, _ in p], [b for _, b in p]) for p in s.values()]
        rhos = [r for r in rhos if np.isfinite(r)]
        return float(np.mean(rhos)) if rhos else float("nan")

    observe = statistique(strates)
    compte = 0
    for _ in range(TIRAGES):
        melange = {m: list(zip(aleatoire.permutation([a for a, _ in p]),
                               [b for _, b in p]))
                   for m, p in strates.items()}
        valeur = statistique(melange)
        if np.isfinite(valeur) and abs(valeur) >= abs(observe):
            compte += 1
    p = (1 + compte) / (1 + TIRAGES)
    detail = {m: spearman([a for a, _ in v], [b for _, b in v])
              for m, v in strates.items()}
    # Une permutation stratifiée n'explore que le produit des factorielles des
    # tailles de strate. Avec deux strates de trois, cela fait 36 arrangements
    # et la plus petite valeur bilatérale vaut 2/36 : aucune donnée, si parfaite
    # soit-elle, ne descend sous alpha. Le dire est le rôle de ce champ.
    arrangements = 1
    for points in strates.values():
        arrangements *= math.factorial(len(points))
    minimum = 2.0 / arrangements
    return {"testable": True, "relation": f"{x} vers {y}",
            "strates": len(strates), "eprouvettes": sum(len(v) for v in strates.values()),
            "rho_stratifie": observe, "p_permutation": p,
            "arrangements_possibles": arrangements,
            "p_minimal_atteignable": minimum,
            "atteignable": bool(minimum <= ALPHA),
            "rho_par_materiau": detail,
            "significatif": bool(p <= ALPHA),
            "concordant": bool(len({np.sign(v) for v in detail.values()}) == 1)}


def main() -> int:
    config = json.loads(SOURCES.read_text(encoding="utf-8"))
    racine = (ICI / config["racine_locale"]).resolve()
    base = base_de("carbures_matrice_acier", racine)
    if base is None:
        base = racine / "carbures_matrice_acier" / "raw"
    if not base.is_dir() or not (base / "Heat treatments.xlsx").exists():
        print("source absente : carbures_matrice_acier")
        print("Résultats commités laissés intacts.")
        return 1

    eprouvettes = extraire(base)
    if len(eprouvettes) < 4:
        print(f"{len(eprouvettes)} éprouvette(s) extraite(s), minimum 4.")
        return 1

    DERIVE.mkdir(exist_ok=True)
    with (DERIVE / "carbures_par_eprouvette.csv").open(
            "w", encoding="utf-8", newline="") as flux:
        redacteur = csv.DictWriter(flux, fieldnames=list(eprouvettes[0]),
                                   lineterminator="\n")
        redacteur.writeheader()
        redacteur.writerows(eprouvettes)

    aleatoire = np.random.default_rng(GRAINE)
    relations = {
        "histoire_vers_trace": tester(eprouvettes, "revenu_C", "durete_HRC", aleatoire),
        "trace_vers_reponse": tester(eprouvettes, "durete_HRC", "tenacite", aleatoire),
        "histoire_vers_reponse": tester(eprouvettes, "revenu_C", "tenacite", aleatoire),
        "histoire_vers_frottement": tester(eprouvettes, "revenu_C",
                                           "cof_haute_charge", aleatoire),
    }

    print(f"{len(eprouvettes)} éprouvettes, "
          f"{len({e['materiau'] for e in eprouvettes})} matériaux.")
    print()
    entete = f"{'relation':<28}{'n':>4}{'rho':>10}{'p':>10}  par matériau"
    print(entete)
    print("-" * len(entete))
    for nom, r in relations.items():
        if not r["testable"]:
            print(f"{nom:<28}{'—':>4}{'—':>10}{'—':>10}  {r['motif']}")
            continue
        detail = "  ".join(f"{m} {v:+.2f}" for m, v in sorted(r["rho_par_materiau"].items()))
        print(f"{nom:<28}{r['eprouvettes']:>4}{r['rho_stratifie']:>10.4f}"
              f"{r['p_permutation']:>10.6f}  {detail}")

    maillons = ("histoire_vers_trace", "trace_vers_reponse")
    parfaits = all(abs(relations[c].get("rho_stratifie", 0)) > 0.99
                   and relations[c].get("concordant") for c in maillons)
    atteignables = all(relations[c].get("atteignable") for c in maillons)
    chaine = all(relations[c].get("significatif") and relations[c].get("concordant")
                 for c in maillons)

    if chaine:
        verdict = "soutient"
        motif = ("les deux maillons sont significatifs et concordants entre "
                 "matériaux")
    elif not atteignables:
        verdict = "indetermine_par_atteignabilite"
        minimums = ", ".join(f"{c} p minimal {relations[c]['p_minimal_atteignable']:.4f}"
                             for c in maillons)
        motif = (f"le plan ne permet pas d'atteindre alpha = {ALPHA} : {minimums}. "
                 f"Les corrélations valent pourtant "
                 f"{relations['histoire_vers_trace']['rho_stratifie']:+.3f} et "
                 f"{relations['trace_vers_reponse']['rho_stratifie']:+.3f}, "
                 f"{'concordantes' if parfaits else 'non concordantes'} entre les deux "
                 f"matériaux. Ce n'est pas un résultat négatif.")
    else:
        verdict = "ne_soutient_pas"
        motif = ("au moins un maillon n'est pas significatif ou change de signe "
                 "selon le matériau")

    print()
    print(f"VERDICT chaîne complète : {verdict}")
    print(f"  {motif}")

    sortie = DERIVE / "RESULTAT_CARBURES.json"
    with sortie.open("w", encoding="utf-8", newline="") as flux:
        flux.write(json.dumps({
            "campagne": "WP-MAT-MEM-2026",
            "famille": "transition_de_phase",
            "jeu": "aciers à outils revenus, 6 éprouvettes, 2 matériaux",
            "histoire": "température du second revenu",
            "trace": "dureté HRC mesurée avant les essais de réponse",
            "reponse": "ténacité à la rupture et coefficient de frottement",
            "alpha": ALPHA, "graine": GRAINE, "tirages": TIRAGES,
            "eprouvettes": len(eprouvettes),
            "relations": relations,
            "chaine_complete": chaine,
            "chaine_de_signe_parfaite_et_concordante": parfaits,
            "atteignable_a_alpha": atteignables,
            "verdict": verdict, "motif": motif,
            "statut_epistemique": "mesures d'instrument sur éprouvettes physiques",
        }, ensure_ascii=False, indent=2) + "\n")
    print(f"écrit : {sortie.relative_to(ICI.parents[1]).as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
