#!/usr/bin/env python3
"""Décrit le contenu d'un jeu téléchargé : volumes, natures de fichiers,
conditions candidates, indices d'histoire, trace, réponse, témoin,
ablation. Écrit INSPECTION.json.

    python inspecter_jeu.py --cle CLE | --toutes
"""
from __future__ import annotations

import argparse
import json
import re
import tempfile
import zipfile
from collections import Counter
from pathlib import Path

ICI = Path(__file__).resolve().parent
SOURCES = ICI / "SOURCES.json"

# Vocabulaire de reconnaissance. Volontairement large : l'inspecteur signale des
# candidats, il ne tranche pas. Un faux positif ici coûte une vérification ; un
# faux négatif coûte un jeu perdu.
MOTIFS = {
    "histoire": (r"ecap|as.?built|lta|anneal|recuit|aged|vieilli|precondition|"
                 r"pretreat|prior|cycle|pass|deform|strain|bias|bdc|history|"
                 r"traitement|trempe|quench|soak|hold"),
    "trace": (r"hardness|durete|hv|ebsd|xrd|hexrd|xps|stm|stem|saed|tem|lattice|"
              r"maille|phase|grain|remanen|coerciv|enthalp|dsc|resistance|"
              r"dilat|microstructure|texture|composition"),
    "reponse": (r"stress|strain|compress|tensile|traction|loss|perte|loop|boucle|"
                r"ratchet|relax|oer|activity|activite|current|courant|"
                r"conversion|selectivity|transformation|kinetic"),
    "temoin": (r"reference|ref\b|control|temoin|blank|vierge|pristine|initial|"
               r"as.?received|untreated|non.?traite|baseline"),
    "ablation": (r"demagnet|degauss|restoration|restaur|solution.?treat|"
                 r"remise.?en.?solution|rejuven|reset|erase|effac|recuit.?de.?restaur"),
    "serie_temporelle": (r"time|temps|_t_|cycle|step|sweep|kinetic|transient|"
                         r"evolution|profile|curve|courbe|loop"),
}

TABULAIRE = {".csv", ".txt", ".dat", ".tsv", ".xlsx", ".xls", ".xlsm", ".asc", ".ctf", ".ang"}
IMAGE = {".tif", ".tiff", ".png", ".jpg", ".jpeg", ".bmp"}
DOCUMENT = {".pdf", ".docx", ".md"}
CALCUL = {".cif", ".xyz", ".vasp", ".outcar", ".poscar"}


def noms_dans(chemin: Path, profondeur: int = 0) -> list[str]:
    """Chemins internes d'une archive, récursivement, ou le chemin lui-même.

    Trois formats doivent être ouverts, faute de quoi des jeux entiers restent
    invisibles.

    `.zip` — y compris **imbriqués** : `ti_nb_sn` livre une archive par figure,
    et se contenter du premier niveau ne montrerait que quatorze noms de zip.

    `.rar` — deux jeux d'aciers moyen-Mn sont livrés en RAR. Sans `rarfile`, ils
    apparaissaient comme « 1 entrée, 0 tabulaire », alors qu'ils contiennent des
    séries de dilatométrie dont les noms encodent l'histoire thermique complète :
    température, vitesse de chauffe, durée de maintien, vitesse de
    refroidissement. Un jeu de premier plan aurait été écarté pour un défaut
    d'outillage.

    `.xlsx` — les noms de feuilles sont ajoutés comme entrées virtuelles.
    Plusieurs jeux ne livrent que des classeurs, et toute leur structure
    expérimentale est dans les onglets.
    """
    suffixe = chemin.suffix.lower()
    if profondeur > 2:
        return [chemin.name]

    if suffixe == ".zip":
        try:
            with zipfile.ZipFile(chemin) as archive:
                noms = [n for n in archive.namelist()
                        if not n.endswith("/") and "__MACOSX" not in n]
                interieurs = []
                for nom in noms:
                    if nom.lower().endswith((".zip", ".xlsx", ".xls", ".xlsm")):
                        with tempfile.TemporaryDirectory() as temporaire:
                            extrait = Path(temporaire) / Path(nom).name
                            extrait.write_bytes(archive.read(nom))
                            interieurs += [f"{nom}/{i}"
                                           for i in noms_dans(extrait, profondeur + 1)]
                return noms + interieurs
        except (zipfile.BadZipFile, OSError, RuntimeError):
            return []

    if suffixe == ".rar":
        try:
            import rarfile
        except ImportError:
            return [f"{chemin.name}  [RAR non ouvert : installer rarfile]"]
        try:
            with rarfile.RarFile(chemin) as archive:
                return [n for n in archive.namelist() if not n.endswith("/")]
        except Exception:
            return [f"{chemin.name}  [RAR illisible]"]

    if suffixe in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
            classeur = load_workbook(chemin, read_only=True)
            feuilles = [f"{chemin.name}#{f}" for f in classeur.sheetnames]
            classeur.close()
            return [chemin.name] + feuilles
        except Exception:
            return [chemin.name]

    return [chemin.name]


def classer(noms: list[str]) -> dict:
    trouve = {cle: Counter() for cle in MOTIFS}
    extensions = Counter()
    for nom in noms:
        minuscule = nom.lower()
        extensions[Path(nom).suffix.lower()] += 1
        for cle, motif in MOTIFS.items():
            for correspondance in set(re.findall(motif, minuscule)):
                trouve[cle][correspondance] += 1
    return {"indices": trouve, "extensions": extensions}


def segments_candidats(noms: list[str]) -> Counter:
    """Segments de chemin susceptibles d'identifier une condition ou un échantillon.

    Un dossier qui se répète à un même niveau avec des noms différents est le
    signe d'une conception factorielle : `As-built/`, `LTA_280/`, `LTA_300/`.
    C'est le meilleur indice automatique de la présence de plusieurs histoires.
    """
    compte = Counter()
    for nom in noms:
        parts = [p for p in Path(nom).parts[:-1] if p not in (".", "..")]
        for profondeur, part in enumerate(parts):
            compte[f"niveau{profondeur}:{part}"] += 1
    return compte


def inspecter(cle: str, racine: Path, source: dict) -> dict:
    brut = racine / cle / "raw"
    if not brut.is_dir():
        return {"cle": cle, "statut": "non téléchargé"}

    fichiers = sorted(p for p in brut.rglob("*") if p.is_file())
    tous_noms: list[str] = []
    for fichier in fichiers:
        tous_noms.extend(noms_dans(fichier))

    analyse = classer(tous_noms)
    segments = segments_candidats(tous_noms)
    extensions = analyse["extensions"]

    tabulaires = sum(n for e, n in extensions.items() if e in TABULAIRE)
    images = sum(n for e, n in extensions.items() if e in IMAGE)
    documents = sum(n for e, n in extensions.items() if e in DOCUMENT)
    calculs = sum(n for e, n in extensions.items() if e in CALCUL)

    # Dossiers répétés au même niveau : candidats « conditions expérimentales ».
    par_niveau: dict[int, Counter] = {}
    for etiquette, compte in segments.items():
        niveau = int(etiquette.split(":", 1)[0].removeprefix("niveau"))
        par_niveau.setdefault(niveau, Counter())[etiquette.split(":", 1)[1]] = compte
    conditions = {}
    for niveau, noms in sorted(par_niveau.items()):
        if 2 <= len(noms) <= 30:
            conditions[f"niveau_{niveau}"] = dict(noms.most_common(30))

    def resume(cle_motif: str) -> dict:
        return dict(analyse["indices"][cle_motif].most_common(12))

    return {
        "cle": cle,
        "famille_declaree": source.get("famille"),
        "statut": "inspecté",
        "volume": {
            "fichiers_telecharges": len(fichiers),
            "octets": sum(f.stat().st_size for f in fichiers),
            "entrees_internes": len(tous_noms),
        },
        "nature_des_fichiers": {
            "tabulaires": tabulaires,
            "images": images,
            "documents": documents,
            "fichiers_de_calcul": calculs,
            "extensions": dict(extensions.most_common(20)),
        },
        "conditions_candidates": conditions,
        "indices": {
            "histoire": resume("histoire"),
            "trace": resume("trace"),
            "reponse": resume("reponse"),
            "temoin": resume("temoin"),
            "ablation": resume("ablation"),
            "serie_temporelle": resume("serie_temporelle"),
        },
        "avertissement_unites": (
            "Le nombre d'entrées internes n'est PAS un nombre d'échantillons "
            "physiques. Dix mille points relevés sur le même morceau de métal "
            "restent une unité expérimentale. Le nombre d'échantillons se lit "
            "dans la source, jamais dans l'arborescence."
        ),
        "suite": (
            "Renseigner une fiche depuis la source elle-même, puis "
            "admettre_jeu.py. L'inspection propose, elle ne décide pas."
        ),
    }


def afficher(rapport: dict) -> None:
    if rapport["statut"] != "inspecté":
        print(f"  {rapport['cle']:<32} {rapport['statut']}")
        return
    v = rapport["volume"]
    n = rapport["nature_des_fichiers"]
    print(f"  {rapport['cle']:<32} {v['octets'] / 1e6:>8.1f} Mo  "
          f"{v['fichiers_telecharges']} fichier(s), {v['entrees_internes']} entrées")
    print(f"      tabulaires {n['tabulaires']}, images {n['images']}, "
          f"documents {n['documents']}, calcul {n['fichiers_de_calcul']}")
    for niveau, noms in rapport["conditions_candidates"].items():
        print(f"      conditions {niveau} : {', '.join(list(noms)[:8])}")
    for cle in ("histoire", "trace", "reponse", "temoin", "ablation"):
        indices = rapport["indices"][cle]
        if indices:
            print(f"      {cle:<16} {', '.join(list(indices)[:8])}")
        elif cle in ("ablation", "temoin"):
            print(f"      {cle:<16} aucun indice")


def main() -> int:
    analyseur = argparse.ArgumentParser(description=__doc__)
    analyseur.add_argument("--cle", action="append")
    analyseur.add_argument("--toutes", action="store_true")
    arguments = analyseur.parse_args()

    config = json.loads(SOURCES.read_text(encoding="utf-8"))
    racine = (ICI / config["racine_locale"]).resolve()
    sources = config["sources"]
    if arguments.cle:
        sources = [s for s in sources if s["cle"] in set(arguments.cle)]
    elif not arguments.toutes:
        analyseur.error("préciser --cle ou --toutes")

    print(f"Racine : {racine}")
    print()
    rapports = []
    for source in sources:
        rapport = inspecter(source["cle"], racine, source)
        rapports.append(rapport)
        afficher(rapport)

    sortie = ICI / "INSPECTION.json"
    with sortie.open("w", encoding="utf-8", newline="") as flux:
        flux.write(json.dumps(
            {"objet": "description automatique des jeux téléchargés ; "
                      "ne conclut rien et n'admet rien",
             "jeux": rapports}, ensure_ascii=False, indent=2) + "\n")
    print()
    print(f"écrit : {sortie.relative_to(ICI.parents[1]).as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
