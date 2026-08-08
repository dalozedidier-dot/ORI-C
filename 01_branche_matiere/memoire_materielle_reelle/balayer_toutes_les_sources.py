#!/usr/bin/env python3
"""Balayage uniforme des sources sans extracteur dédié.

Dose : nombre suivi d'une unité dans un composant du chemin, notation
1p25 comprise. Clé de groupe arrêtée au composant porteur.
Réponse : moyenne de la dernière colonne numérique.
Témoin : permutation des doses dans chaque série.
Écrit BALAYAGE_TOUTES_SOURCES.json.

    python balayer_toutes_les_sources.py
"""
from __future__ import annotations

import csv
import json
import re
import warnings
from collections import defaultdict
from pathlib import Path

import numpy as np

warnings.filterwarnings("ignore")

ICI = Path(__file__).resolve().parent
SOURCES = ICI / "SOURCES.json"
DERIVE = ICI / "derive"

ALPHA = 0.05
GRAINE = 20260809
TIRAGES = 10000
DOSES_MINIMUM = 3
GROUPES_MINIMUM = 3

# Sources déjà traitées par un extracteur dédié : le balayage les ignore pour ne
# pas produire un second verdict, moins soigné, sur les mêmes données.
DEJA_TRAITEES = {"fabest_lcf", "medium_mn_a", "recuit_thermique_polymere"}

UNITES = (r"(?:mT|T|MPa|GPa|kPa|Hz|kHz|mA|mV|V|W|°?C|K|days?|jours?|hs?|min|s|pct|%|cycles?|passes?|Gy|nm|um|µm)")
JETON = re.compile(rf"(?<![A-Za-z0-9])(\d+(?:[.,p]\d+)?)\s*({UNITES})(?![A-Za-z])", re.I)
# Motif `température-durée` très répandu : 100-12, 90-5.
COUPLE = re.compile(r"(?<![A-Za-z0-9])(\d{2,3})-(\d{1,3})(?![A-Za-z0-9])")

TABULAIRE = {".csv", ".txt", ".dat", ".tsv", ".asc", ".xlsx", ".xls", ".xlsm"}


def spearman(x, y) -> float:
    x, y = np.asarray(x, float), np.asarray(y, float)
    if x.size < 3:
        return float("nan")
    def rangs(v):
        o = np.argsort(v, kind="mergesort")
        r = np.empty(len(v), dtype=float)
        r[o] = np.arange(len(v), dtype=float)
        return r
    rx, ry = rangs(x), rangs(y)
    if rx.std() == 0 or ry.std() == 0:
        return float("nan")
    return float(np.corrcoef(rx, ry)[0, 1])


def p_sign_flip(valeurs: np.ndarray, aleatoire) -> tuple[float, str]:
    n = valeurs.size
    if n == 0:
        return 1.0, "aucune valeur"
    observe = abs(float(valeurs.mean()))
    if n <= 20:
        compte = 0
        for masque in range(1 << n):
            signes = np.array([1.0 if masque >> i & 1 else -1.0 for i in range(n)])
            if abs(float((valeurs * signes).mean())) >= observe:
                compte += 1
        return compte / (1 << n), f"sign-flip exact, {1 << n} attributions"
    compte = sum(1 for _ in range(TIRAGES)
                 if abs(float((valeurs * aleatoire.choice((-1.0, 1.0), size=n)).mean()))
                 >= observe)
    return (1 + compte) / (1 + TIRAGES), f"sign-flip Monte-Carlo, {TIRAGES} tirages"


def dose_et_groupe(chemin: Path, base: Path) -> tuple[float, str] | None:
    """Dose lue dans le chemin, et clé de groupe arrêtée au composant qui la porte.

    Le composant est le dossier ou le nom de fichier où le jeton apparaît. Tout
    ce qui suit — horodatage, numéro de réplicat, nom d'appareil — est écarté de
    la clé : ces variations ne définissent pas une série dose-réponse, elles la
    peuplent.
    """
    parties = list(chemin.relative_to(base).parts)
    for rang, composant in enumerate(parties):
        couple = COUPLE.search(composant)
        if couple:
            dose = float(couple.group(2))
            marque = composant[:couple.start()] + "{T}-{d}" + composant[couple.end():]
            return dose, "/".join(parties[:rang] + [marque])
        jeton = JETON.search(composant)
        if jeton:
            dose = float(jeton.group(1).replace(",", ".").replace("p", "."))
            marque = composant[:jeton.start()] + "{dose}" + composant[jeton.end():]
            return dose, "/".join(parties[:rang] + [marque])
    return None


def reponse(chemin: Path) -> float:
    """Moyenne de la dernière colonne numérique. Définition uniforme, déclarée."""
    suffixe = chemin.suffix.lower()
    try:
        if suffixe in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            classeur = load_workbook(chemin, read_only=True, data_only=True)
            feuille = classeur[classeur.sheetnames[0]]
            colonnes = defaultdict(list)
            for ligne in feuille.iter_rows(values_only=True):
                for indice, valeur in enumerate(ligne or ()):
                    if isinstance(valeur, (int, float)):
                        colonnes[indice].append(float(valeur))
            classeur.close()
        else:
            colonnes = defaultdict(list)
            with chemin.open(encoding="utf-8", errors="replace") as flux:
                for brut in flux:
                    ligne = brut.strip()
                    if not ligne or ligne.startswith(("#", "%", "!")):
                        continue
                    for separateur in (";", "\t", ",", None):
                        parties = ligne.split(separateur) if separateur else ligne.split()
                        if len(parties) >= 2:
                            break
                    for indice, morceau in enumerate(parties):
                        try:
                            colonnes[indice].append(float(morceau.replace(",", ".")))
                        except ValueError:
                            continue
    except Exception:
        return float("nan")
    utiles = [i for i, v in sorted(colonnes.items()) if len(v) >= 5]
    if not utiles:
        return float("nan")
    valeurs = colonnes[utiles[-1]]
    return float(np.mean(valeurs))


def balayer(cle: str, base: Path, aleatoire) -> dict:
    fichiers = [p for p in base.rglob("*")
                if p.is_file() and p.suffix.lower() in TABULAIRE]
    groupes: dict[str, list[tuple[float, float]]] = defaultdict(list)
    for chemin in fichiers:
        trouve = dose_et_groupe(chemin, base)
        if not trouve:
            continue
        dose, groupe = trouve
        valeur = reponse(chemin)
        if np.isfinite(valeur):
            groupes[groupe].append((dose, valeur))

    rhos, rhos_permutes, detail = [], [], {}
    for groupe, points in sorted(groupes.items()):
        # Plusieurs fichiers peuvent partager une dose : on moyenne, une valeur
        # par dose. Sans quoi les réplicats pèseraient comme des doses.
        par_dose: dict[float, list[float]] = defaultdict(list)
        for dose, valeur in points:
            par_dose[dose].append(valeur)
        if len(par_dose) < DOSES_MINIMUM:
            continue
        doses = sorted(par_dose)
        valeurs = [float(np.mean(par_dose[d])) for d in doses]
        rho = spearman(doses, valeurs)
        rho_permute = spearman(aleatoire.permutation(doses), valeurs)
        if np.isfinite(rho):
            rhos.append(rho)
            detail[groupe[:70]] = {"fichiers": len(points),
                                   "doses_distinctes": len(doses), "rho": rho}
        if np.isfinite(rho_permute):
            rhos_permutes.append(rho_permute)

    rhos = np.array(rhos)
    rhos_permutes = np.array(rhos_permutes)
    if rhos.size < GROUPES_MINIMUM:
        return {"cle": cle, "fichiers_tabulaires": len(fichiers),
                "groupes_dose_reponse": int(rhos.size),
                "verdict": "aucune_structure_dose_reponse_reperee",
                "motif": (f"{rhos.size} série(s) dose-réponse trouvée(s), "
                          f"minimum {GROUPES_MINIMUM}")}

    p, methode = p_sign_flip(rhos, aleatoire)
    p_permute, _ = p_sign_flip(rhos_permutes, aleatoire)
    minimum = 2.0 / 2 ** rhos.size if rhos.size <= 20 else 1.0 / (1 + TIRAGES)

    if p_permute <= ALPHA:
        verdict = "invalide"
        motif = (f"le témoin permuté est significatif, p = {p_permute:.4f} : la "
                 f"corrélation ne vient pas de l'ordre des doses")
    elif minimum > ALPHA:
        verdict = "indetermine_par_atteignabilite"
        motif = f"{rhos.size} séries, p minimal {minimum:.4f} > {ALPHA}"
    elif p <= ALPHA:
        verdict = "soutient"
        motif = "la réponse suit la dose d'histoire, le témoin permuté ne suit pas"
    else:
        verdict = "ne_soutient_pas"
        motif = f"p = {p:.4f} > {ALPHA}"

    return {"cle": cle, "fichiers_tabulaires": len(fichiers),
            "groupes_dose_reponse": int(rhos.size),
            "rho_moyen": float(rhos.mean()),
            "fraction_negative": float((rhos < 0).mean()),
            "p": p, "methode_du_temoin": methode,
            "p_minimal_atteignable": minimum,
            "temoin_permute": {"rho_moyen": float(rhos_permutes.mean())
                               if rhos_permutes.size else float("nan"),
                               "p": p_permute},
            "series": dict(list(detail.items())[:12]),
            "verdict": verdict, "motif": motif}


def main() -> int:
    config = json.loads(SOURCES.read_text(encoding="utf-8"))
    racine = (ICI / config["racine_locale"]).resolve()
    aleatoire = np.random.default_rng(GRAINE)

    a_traiter = [s for s in config["sources"]
                 if not s["cle"].startswith("iodp_")
                 and s["cle"] not in DEJA_TRAITEES
                 and (racine / s["cle"] / "exploitable").is_dir()]

    print(f"{len(a_traiter)} source(s) balayée(s), procédure identique pour toutes.")
    print(f"Réponse : moyenne de la dernière colonne numérique. Témoin : permutation "
          f"des doses dans chaque série.")
    print()
    entete = (f"{'source':<34}{'fich.':>6}{'séries':>8}{'rho moyen':>11}"
              f"{'p':>9}{'p min':>9}   verdict")
    print(entete)
    print("-" * len(entete))

    rapports = []
    for source in a_traiter:
        rapport = balayer(source["cle"], racine / source["cle"] / "exploitable", aleatoire)
        rapport["famille"] = source["famille"]
        rapports.append(rapport)
        if "rho_moyen" in rapport:
            print(f"{rapport['cle']:<34}{rapport['fichiers_tabulaires']:>6}"
                  f"{rapport['groupes_dose_reponse']:>8}{rapport['rho_moyen']:>11.4f}"
                  f"{rapport['p']:>9.4f}{rapport['p_minimal_atteignable']:>9.4f}"
                  f"   {rapport['verdict']}")
        else:
            print(f"{rapport['cle']:<34}{rapport['fichiers_tabulaires']:>6}"
                  f"{rapport['groupes_dose_reponse']:>8}{'—':>11}{'—':>9}{'—':>9}"
                  f"   {rapport['verdict']}")

    soutiennent = [r for r in rapports if r["verdict"] == "soutient"]
    familles = {r["famille"] for r in soutiennent}
    print()
    print(f"{len(soutiennent)} source(s) soutiennent, "
          f"{len(familles)} famille(s) : {', '.join(sorted(familles)) or 'aucune'}")

    DERIVE.mkdir(exist_ok=True)
    sortie = DERIVE / "BALAYAGE_TOUTES_SOURCES.json"
    with sortie.open("w", encoding="utf-8", newline="") as flux:
        flux.write(json.dumps({
            "campagne": "WP-MAT-MEM-2026",
            "objet": ("balayage uniforme des sources sans extracteur dédié. Dit "
                      "lesquelles portent une structure dose-réponse repérable ; "
                      "ne remplace pas un extracteur propre."),
            "definition_de_la_reponse": "moyenne de la dernière colonne numérique",
            "temoin": "permutation des doses à l'intérieur de chaque série",
            "alpha": ALPHA, "graine": GRAINE,
            "doses_minimum_par_serie": DOSES_MINIMUM,
            "series_minimum_par_source": GROUPES_MINIMUM,
            "sources_a_extracteur_dedie": sorted(DEJA_TRAITEES | {"iodp_*"}),
            "sources": rapports,
        }, ensure_ascii=False, indent=2) + "\n")
    print(f"écrit : {sortie.relative_to(ICI.parents[1]).as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
