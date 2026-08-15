"""Extraction quantitative du classeur Nader AVT, sans extrapolation causale."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[2]
SOURCE = ROOT / "donnees_externes/nader_vesicules_2026/S. Nader AVT raw data.xlsx"
PROTOCOL = ROOT / "03_branche_vivant/lignees_vesicules/PROTOCOLE_PACC_CAUSAL_PROSPECTIF_V1.json"
OUT = HERE / "resultats"
EMPIRICAL_SHEETS = ["Fig 1A & S9", "Fig1B", "Fig1C", "FigS2", "FigS4", "FigS5", "FigS7", "FigS10", "FigS11", "FigS12", "FigS13", "FigS15"]
CHROMATOGRAM_SHEETS = ["Fig 1A & S9", "Fig1B", "Fig1C", "FigS5", "FigS10", "FigS11", "FigS12", "FigS13"]


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")


def all_numeric_cells(book: openpyxl.Workbook) -> pd.DataFrame:
    rows = []
    for sheet_name in EMPIRICAL_SHEETS:
        sheet = book[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    rows.append({"sheet": sheet_name, "row": cell.row, "column": cell.column, "value": float(cell.value)})
    return pd.DataFrame(rows)


def chromatograms(book: openpyxl.Workbook) -> pd.DataFrame:
    records = []
    for sheet_name in CHROMATOGRAM_SHEETS:
        sheet = book[sheet_name]
        context = str(sheet.cell(1, 1).value or "")
        for row in range(1, sheet.max_row + 1):
            first = sheet.cell(row, 1).value
            sample = sheet.cell(row, 2).value
            if isinstance(first, str) and first.strip() and "Sample name" not in first:
                context = first.strip()
            values = [sheet.cell(row, col).value for col in range(3, 49)]
            numeric_count = sum(isinstance(v, (int, float)) for v in values)
            if not isinstance(sample, str) or "_" not in sample or numeric_count < 10:
                continue
            condition, replicate = sample.rsplit("_", 1)
            concentration = None
            if re.fullmatch(r"-?\d+(?:\.\d+)?", condition):
                concentration = float(condition)
            for fraction, value in enumerate(values, 1):
                if isinstance(value, (int, float)):
                    records.append(
                        {
                            "sheet": sheet_name,
                            "context": context,
                            "landing_solution": "2:1 decanoic acid:decanol" if ("decanoic" in context.lower() or sheet_name == "Fig1C") else "not_explicit_in_block",
                            "condition": condition,
                            "concentration_mM_if_encoded": concentration,
                            "replicate": replicate,
                            "elution_fraction": fraction,
                            "fluorescence": float(value),
                        }
                    )
    return pd.DataFrame(records)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    book = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    numeric = all_numeric_cells(book)
    chrom = chromatograms(book)
    sheet = book["FigS15"]
    sizes = [float(sheet.cell(3, col).value) for col in range(2, sheet.max_column + 1) if isinstance(sheet.cell(3, col).value, (int, float))]
    size_table = pd.DataFrame({"particle_size_um": sizes, "particle_size_nm": np.asarray(sizes) * 1000})
    protocol = json.loads(PROTOCOL.read_text(encoding="utf-8"))
    # Formulation canonique locale : ne pas propager un ancien mojibake du
    # protocole dans cet artefact dérivé.
    target = "100 nm extrusion; required population median Z-average 80-150 nm and median PDI <= 0.25"
    two_to_one = chrom.loc[chrom.landing_solution == "2:1 decanoic acid:decanol"].copy()
    concentration_values = sorted(two_to_one.concentration_mM_if_encoded.dropna().unique().tolist())
    result = {
        "schema": "oric.nader-avt-real-characterisation.v1",
        "status": "real_physical_characterisation_only",
        "source_sha256": sha256(SOURCE),
        "quantitative_extraction": {
            "n_empirical_numeric_cells": len(numeric),
            "excluded_sheet": "FigS6 (theoretical travel-time calculation; not empirical data)",
            "n_chromatogram_measurements": len(chrom),
            "n_2_to_1_decanoic_acid_decanol_measurements": len(two_to_one),
            "encoded_concentration_range_mM": [float(min(concentration_values)), float(max(concentration_values))] if concentration_values else None,
            "encoded_concentrations_mM": concentration_values,
        },
        "observed_particle_sizes": {
            "n_particles": len(sizes),
            "min_nm": float(np.min(size_table.particle_size_nm)),
            "median_nm": float(np.median(size_table.particle_size_nm)),
            "q025_nm": float(np.quantile(size_table.particle_size_nm, 0.025)),
            "q975_nm": float(np.quantile(size_table.particle_size_nm, 0.975)),
            "max_nm": float(np.max(size_table.particle_size_nm)),
        },
        "comparison_with_VES_PACC_INT_01": {
            "protocol_target": target,
            "fraction_observed_particles_between_80_and_150_nm": float(size_table.particle_size_nm.between(80, 150).mean()),
            "direct_comparability": False,
            "reason": "Nader reports transformed-particle microscopy sizes; VES-PACC-INT-01 targets a population DLS Z-average and PDI after a 100 nm extrusion operator",
        },
        "fields_requested_but_not_paired_in_workbook": {
            "pH": "not found as a sample-level quantitative field",
            "time_resolved_stability": "not paired to the chromatogram samples",
            "DLS_PDI": "not reported",
        },
        "qualification": {
            "real_data": True,
            "synthetic_or_simulated_scientific_data": False,
            "may_replace_control_do_m_sham": False,
            "use": "physical calibration of observed vesicle formation and size ranges only",
        },
    }
    numeric.to_csv(OUT / "NADER_ALL_EMPIRICAL_NUMERIC_CELLS.csv", index=False, lineterminator="\n")
    chrom.to_csv(OUT / "NADER_CHROMATOGRAMS_LONG.csv", index=False, lineterminator="\n")
    two_to_one.to_csv(OUT / "NADER_DECANOIC_ACID_DECANOL_2_TO_1.csv", index=False, lineterminator="\n")
    size_table.to_csv(OUT / "NADER_PARTICLE_SIZES.csv", index=False, lineterminator="\n")
    write_json(OUT / "RESULTAT_NADER_2026.json", result)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
