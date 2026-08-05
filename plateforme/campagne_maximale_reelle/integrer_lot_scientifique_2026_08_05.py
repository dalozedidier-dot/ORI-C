"""Intègre le lot scientifique reçu le 5 août 2026.

Aucune valeur absente n'est inventée. Les tables canoniques ne sont produites
que lorsque le fichier source contient effectivement les variables requises.
Les sources utiles mais incompatibles sont converties en tables auxiliaires et
restent exclues des verdicts de la campagne.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import tarfile
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.io import netcdf_file

ROOT = Path(__file__).resolve().parents[2]
HERE = Path(__file__).resolve().parent
RAW = ROOT / "donnees_externes/lot_scientifique_maximal_2026_08_05/raw"

SOURCE_URLS = {
    "KeySeries.zip": "https://data.giss.nasa.gov/gistemp/uncertainty/v2.0/KeySeries.zip",
    "HadCRUT.5.1.0.0.analysis.ensemble_series.global.monthly.csv": "https://www.metoffice.gov.uk/hadobs/hadcrut5/data/HadCRUT.5.1.0.0/analysis/diagnostics/HadCRUT.5.1.0.0.analysis.ensemble_series.global.monthly.csv",
    "CMIP6_NASST_historical-ssp585-ssp245.zip": "https://zenodo.org/records/13853454/files/CMIP6_NASST_historical-ssp585-ssp245.zip?download=1",
    "original_ensembles.tar.gz": "https://zenodo.org/records/10137680/files/original_ensembles.tar.gz?download=1",
    "globalmean.zip": "https://zenodo.org/records/13760666/files/globalmean.zip?download=1",
    "kida.uva.2024 (1).zip": "https://kida.astrochem-tools.org/uploads/models/kida.uva.2024.zip",
    "rate22_final.rates": "https://umistdatabase.uk/files/rate22_final.rates",
    "rate22_revised_CtoO_0.44.specs": "https://umistdatabase.uk/files/rate22_revised_CtoO_0.44.specs",
    "rate22_dipole.specs": "https://umistdatabase.uk/files/rate22_dipole.specs",
    "SIMPLE_CCSNe_v3p1.hdf5": "https://zenodo.org/records/19063105/files/SIMPLE_CCSNe_v3p1.hdf5?download=1",
    "Ca isotope data for analyzed lunar samples_Fu et al 2023.xlsx": "https://zenodo.org/records/8197851/files/Ca%20isotope%20data%20for%20analyzed%20lunar%20samples_Fu%20et%20al%202023.xlsx?download=1",
    "Rodriguez_Supplementary_TableS1.xlsx": "https://zenodo.org/records/17179896/files/Rodriguez_Supplementary_TableS1.xlsx?download=1",
    "TableS3.xlsx": "https://zenodo.org/records/10958755/files/TableS3.xlsx?download=1",
    "TableS4.xlsx": "https://zenodo.org/records/10958755/files/TableS4.xlsx?download=1",
    "Dataset_Fig4_FigS1_FigS2_DC_1.csv": "https://zenodo.org/records/17357322/files/Dataset_Fig4_FigS1_FigS2_DC_1.csv?download=1",
    "Dataset_Fig5_bulk_core.xlsx": "https://zenodo.org/records/17357322/files/Dataset_Fig5_bulk_core.xlsx?download=1",
    "Dataset_Fig2.xlsx": "https://zenodo.org/records/17357322/files/Dataset_Fig2.xlsx?download=1",
    "Murchison-vacc-Zn-GAS.csv": "https://zenodo.org/records/8304787/files/Murchison-vacc-Zn-GAS.csv?download=1",
    "Murch-50airZn-GAS.csv": "https://zenodo.org/records/8304787/files/Murch-50airZn-GAS.csv?download=1",
    "Murch-100airZn-GAS.csv": "https://zenodo.org/records/8304787/files/Murch-100airZn-GAS.csv?download=1",
    "210210_MurchisonSteppedHeating_ICPMSData.csv": "https://zenodo.org/records/8304787/files/210210_MurchisonSteppedHeating_ICPMSData.csv?download=1",
    "Measurement_data.csv": "https://zenodo.org/records/3588334/files/Measurement_data.csv?download=1",
    "HMM_summarized_results.zip": "https://zenodo.org/records/10780716/files/HMM_summarized_results.zip?download=1",
    "aphid_PGN_count.csv": "https://zenodo.org/records/5517159/files/aphid_PGN_count.csv?download=1",
    "AA_abundances.csv": "https://zenodo.org/records/18538333/files/AA_abundances.csv?download=1",
    "AA_uncertainties.csv": "https://zenodo.org/records/18538333/files/AA_uncertainties.csv?download=1",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, lineterminator="\n")


def numeric(value: Any) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def annual_mean_rows(
    values: np.ndarray,
    years: np.ndarray,
    *,
    models: list[str],
    scenario: str,
    variable: str,
    region: str,
    members: list[str] | None = None,
) -> list[dict[str, Any]]:
    values = np.asarray(values, dtype=float)
    years = np.asarray(years, dtype=int)
    if values.ndim == 1:
        values = values[None, :]
    rows: list[dict[str, Any]] = []
    unique_years = np.unique(years)
    for index in range(values.shape[0]):
        model = models[index] if len(models) == values.shape[0] else models[0]
        member = members[index] if members and len(members) == values.shape[0] else str(index + 1)
        series = values[index]
        for year in unique_years:
            selected = series[years == year]
            selected = selected[np.isfinite(selected) & (selected > -9000)]
            if selected.size:
                rows.append({
                    "model": model,
                    "scenario": scenario,
                    "member": member,
                    "time": int(year),
                    "variable": variable,
                    "value": float(np.mean(selected)),
                    "region": region,
                })
    return rows


def build_climate_ensemble(data_dir: Path) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    source_counts: dict[str, int] = {}

    # NASA GISTEMP : ensemble d'incertitude observationnelle, moyenne globale.
    key_zip = RAW / "KeySeries.zip"
    with zipfile.ZipFile(key_zip) as archive, tempfile.TemporaryDirectory(prefix="oric-gistemp-") as td:
        member = "KeySeries/ensembleCombinedSeries_Global.nc"
        archive.extract(member, td)
        with h5py.File(Path(td) / member, "r") as handle:
            values = np.asarray(handle["tas"], dtype=float)
            days = np.asarray(handle["time"], dtype=float)
            dates = pd.Timestamp("1880-01-01") + pd.to_timedelta(days, unit="D")
            block = annual_mean_rows(
                values,
                dates.year.to_numpy(),
                models=["NASA_GISTEMP_observation"],
                scenario="observational_uncertainty",
                variable="temperature_anomaly_C",
                region="global",
                members=[f"ens_{i + 1:03d}" for i in range(values.shape[0])],
            )
            rows.extend(block)
            source_counts["NASA_GISTEMP_KeySeries"] = len(block)

    # HadCRUT5 : second ensemble observationnel indépendant.
    had_path = RAW / "HadCRUT.5.1.0.0.analysis.ensemble_series.global.monthly.csv"
    had = pd.read_csv(had_path)
    had["year"] = pd.to_datetime(had["Time"], format="%Y-%m").dt.year
    realization_cols = [column for column in had.columns if column.startswith("Realization ")]
    block_rows: list[dict[str, Any]] = []
    for column in realization_cols:
        grouped = pd.to_numeric(had[column], errors="coerce").groupby(had["year"]).mean()
        member = column.replace("Realization ", "ens_").zfill(7)
        for year, value in grouped.items():
            if pd.notna(value):
                block_rows.append({
                    "model": "HadCRUT5_observation",
                    "scenario": "observational_uncertainty",
                    "member": member,
                    "time": int(year),
                    "variable": "temperature_anomaly_C",
                    "value": float(value),
                    "region": "global",
                })
    rows.extend(block_rows)
    source_counts["HadCRUT5_200_members"] = len(block_rows)

    # 200 trajectoires CMIP6 de température de surface de l'Atlantique Nord.
    nasst_zip = RAW / "CMIP6_NASST_historical-ssp585-ssp245.zip"
    block_rows = []
    with zipfile.ZipFile(nasst_zip) as archive, tempfile.TemporaryDirectory(prefix="oric-nasst-") as td:
        for member in sorted(name for name in archive.namelist() if name.endswith(".nc")):
            match = re.match(r"(.+?)_(r\d+)_historical-(ssp\d+)\.nc$", Path(member).name)
            if not match:
                continue
            model, realization, scenario = match.groups()
            archive.extract(member, td)
            with netcdf_file(str(Path(td) / member), "r", mmap=False) as dataset:
                years = np.asarray(dataset.variables["year"].data, dtype=float).ravel()
                values = np.asarray(dataset.variables["NASST"].data, dtype=float).ravel()
            for year, value in zip(years, values):
                if np.isfinite(year) and np.isfinite(value):
                    block_rows.append({
                        "model": model,
                        "scenario": f"historical-{scenario}",
                        "member": realization,
                        "time": int(round(float(year))),
                        "variable": "NASST_anomaly",
                        "value": float(value),
                        "region": "North_Atlantic",
                    })
    rows.extend(block_rows)
    source_counts["CMIP6_NASST"] = len(block_rows)

    # Ensemble multi-modèles CMIP6 global, trois scénarios.
    ensemble_tar = RAW / "original_ensembles.tar.gz"
    block_rows = []
    with tarfile.open(ensemble_tar, "r:gz") as archive, tempfile.TemporaryDirectory(prefix="oric-cmip6-") as td:
        selected = [
            member for member in archive.getmembers()
            if re.search(r"TAS_mon_CMIP6_hist\+ssp(126|245|585)_global\.nc$", member.name)
        ]
        for member in sorted(selected, key=lambda item: item.name):
            archive.extract(member, td, filter="data")
            scenario_match = re.search(r"hist\+(ssp\d+)_global", member.name)
            scenario = f"historical-{scenario_match.group(1)}" if scenario_match else "unknown"
            with h5py.File(Path(td) / member.name, "r") as handle:
                values = np.asarray(handle["TAS"], dtype=float) - 273.15
                models = [item.decode("utf-8") if isinstance(item, bytes) else str(item) for item in handle["model"][:]]
                days = np.asarray(handle["time"], dtype=float)
                dates = pd.Timestamp("1850-01-01") + pd.to_timedelta(days, unit="D")
            block_rows.extend(annual_mean_rows(
                values,
                dates.year.to_numpy(),
                models=models,
                scenario=scenario,
                variable="near_surface_air_temperature_C",
                region="global",
                members=["ensemble_mean"] * len(models),
            ))
    rows.extend(block_rows)
    source_counts["CMIP6_global_TAS"] = len(block_rows)

    # Expériences idéalisées de réponse radiative et thermique.
    global_zip = RAW / "globalmean.zip"
    block_rows = []
    with zipfile.ZipFile(global_zip) as archive, tempfile.TemporaryDirectory(prefix="oric-globalmean-") as td:
        for member in sorted(name for name in archive.namelist() if name.endswith(".nc")):
            archive.extract(member, td)
            filename = Path(member).name
            model_match = re.match(r"RR_toa_([^_]+(?:\.[^_]+)?)_(.+?)\.rk_", filename)
            atmosphere_model = model_match.group(1) if model_match else "unknown_atmosphere_model"
            experiment = model_match.group(2) if model_match else filename.removesuffix(".nc")
            if "4xCO2" in experiment:
                scenario = "idealized_4xCO2"
            elif re.search(r"p4[kK]", experiment):
                scenario = "idealized_plus4K"
            elif re.search(r"p2[kK]", experiment):
                scenario = "idealized_plus2K"
            else:
                scenario = "amip_control"
            sst_source = "unknown"
            source_match = re.search(r"SST_([^\.]+)", experiment)
            if source_match:
                sst_source = source_match.group(1)
            with h5py.File(Path(td) / member, "r") as handle:
                years = np.asarray(handle["year"], dtype=float)
                if "dts_gm" not in handle:
                    continue
                values = np.asarray(handle["dts_gm"], dtype=float)
            for year, value in zip(years, values):
                if np.isfinite(year) and np.isfinite(value):
                    block_rows.append({
                        "model": f"{atmosphere_model}:{sst_source}",
                        "scenario": scenario,
                        "member": filename,
                        "time": int(round(float(year))),
                        "variable": "surface_temperature_response_C",
                        "value": float(value),
                        "region": "global",
                    })
    rows.extend(block_rows)
    source_counts["idealized_radiative_response"] = len(block_rows)

    frame = pd.DataFrame(rows).drop_duplicates(
        ["model", "scenario", "member", "time", "variable", "region"]
    ).sort_values(["variable", "region", "scenario", "model", "member", "time"])
    write_csv(data_dir / "modern_climate_ensemble.csv", frame)
    return {
        "rows": int(len(frame)),
        "models": int(frame["model"].nunique()),
        "scenarios": sorted(frame["scenario"].unique().tolist()),
        "variables": sorted(frame["variable"].unique().tolist()),
        "regions": sorted(frame["region"].unique().tolist()),
        "source_rows": source_counts,
    }


def parse_umist_network() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    path = RAW / "rate22_final.rates"
    with path.open(encoding="utf-8", errors="replace") as stream:
        for line in stream:
            parts = line.rstrip().split(":")
            if len(parts) < 18:
                continue
            reactants = [parts[2].strip(), parts[3].strip()]
            products = [parts[index].strip() for index in range(4, 8)]
            alpha = numeric(parts[9])
            tmin = numeric(parts[12])
            tmax = numeric(parts[13])
            if alpha is None:
                continue
            rows.append({
                "reaction_id": f"UMIST22:{parts[0].strip()}",
                "reactants": "+".join(item for item in reactants if item),
                "products": "+".join(item for item in products if item),
                "rate": alpha,
                "temperature_min": max(tmin or 0.0, 0.0),
                "temperature_max": tmax if tmax and tmax > 0 else 10000.0,
                "source_network": "UMIST_RATE22",
                "reaction_type": parts[1].strip(),
                "beta": numeric(parts[10]),
                "gamma": numeric(parts[11]),
                "uncertainty_factor": np.nan,
                "uncertainty_type": "",
                "reference": parts[16].strip('" '),
            })
    return pd.DataFrame(rows)


def parse_kida_network() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    path = RAW / "kida.uva.2024 (1).zip"
    with zipfile.ZipFile(path) as archive:
        text = archive.read("kida.uva.2024/gas_reactions_kida.uva.2024.in").decode("utf-8", errors="replace")
    for line in text.splitlines()[1:]:
        if len(line) < 160 or line.startswith("!"):
            continue
        reactants = [line[index:index + 11].strip() for index in (0, 11, 22)]
        products = [line[index:index + 11].strip() for index in (34, 45, 56, 67, 78)]
        fields = line[90:].split()
        if len(fields) < 13:
            continue
        alpha = numeric(fields[0])
        if alpha is None:
            continue
        rows.append({
            "reaction_id": f"KIDA2024:{fields[10]}",
            "reactants": "+".join(item for item in reactants if item),
            "products": "+".join(item for item in products if item),
            "rate": alpha,
            "temperature_min": max(numeric(fields[7]) or 0.0, 0.0),
            "temperature_max": numeric(fields[8]) or 10000.0,
            "source_network": "KIDA_UVA_2024",
            "reaction_type": fields[6],
            "beta": numeric(fields[1]),
            "gamma": numeric(fields[2]),
            "uncertainty_factor": numeric(fields[3]),
            "uncertainty_type": fields[5],
            "reference": "KIDA reaction number " + fields[10],
        })
    return pd.DataFrame(rows)


def build_molecular_inventory(data_dir: Path) -> dict[str, Any]:
    """Construit l'inventaire initial réellement fourni avec UMIST Rate22.

    Les abondances d'acides aminés du lot sont conservées dans une table
    auxiliaire séparée : elles ne sont pas un inventaire interstellaire et ne
    doivent pas servir à valider le réseau astro-chimique.
    """
    specs_path = RAW / "rate22_revised_CtoO_0.44.specs"
    rows: list[dict[str, Any]] = []
    section = None
    for raw_line in specs_path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if line == "9999 Conserved:":
            section = "conserved"
            continue
        if line == "9999 Parents:":
            section = "parents"
            continue
        if section == "conserved":
            match = re.match(r"^\d+\s+(\S+)\s+([0-9.Ee+-]+)$", line)
        elif section == "parents":
            match = re.match(r"^(\S+)\s+([0-9.Ee+-]+)$", line)
        else:
            match = None
        if not match:
            continue
        species, abundance = match.groups()
        amount = numeric(abundance)
        if amount is None:
            continue
        rows.append({
            "environment_id": "UMIST_RATE22_CtoO_0.44_initial_conditions",
            "species": species,
            "abundance": amount,
            "uncertainty": np.nan,
            "environment_class": "astrochemical_model_initial_condition",
            "inventory_kind": section,
        })
    frame = pd.DataFrame(rows).drop_duplicates(["environment_id", "species"])
    write_csv(data_dir / "molecular_inventory.csv", frame)

    # Compilation moléculaire reçue : conservée sans la présenter comme ciel observé.
    abundance = pd.read_csv(RAW / "AA_abundances.csv")
    uncertainty = pd.read_csv(RAW / "AA_uncertainties.csv")
    species_col = abundance.columns[0]
    metadata = abundance.iloc[:3].set_index(species_col)
    data = abundance.iloc[3:].copy()
    uncertainty_data = uncertainty.iloc[3:].copy().set_index(uncertainty.columns[0])
    auxiliary_rows: list[dict[str, Any]] = []
    for environment in abundance.columns[1:]:
        environment_class = metadata.at["Sample Class", environment] if "Sample Class" in metadata.index else np.nan
        for species, value in data[[species_col, environment]].itertuples(index=False, name=None):
            amount = numeric(value)
            if amount is None or not str(species).strip():
                continue
            unc = np.nan
            if species in uncertainty_data.index and environment in uncertainty_data.columns:
                candidate = uncertainty_data.at[species, environment]
                if isinstance(candidate, pd.Series):
                    candidate = candidate.iloc[0]
                parsed = numeric(candidate)
                if parsed is not None:
                    unc = parsed
            auxiliary_rows.append({
                "environment_id": environment,
                "species": str(species).strip(),
                "abundance": amount,
                "uncertainty": unc,
                "environment_class": environment_class,
                "inventory_kind": "amino_acid_abundance_compilation",
            })
    auxiliary = pd.DataFrame(auxiliary_rows).drop_duplicates(["environment_id", "species"])
    write_csv(data_dir / "molecular_inventory_amino_acids_auxiliary.csv", auxiliary)
    return {
        "rows": int(len(frame)),
        "environments": int(frame["environment_id"].nunique()),
        "species": int(frame["species"].nunique()),
        "uncertainty_coverage": 0.0,
        "scope": "Conditions initiales du modèle UMIST Rate22, pas un inventaire observationnel du ciel.",
        "amino_acid_auxiliary_rows": int(len(auxiliary)),
        "amino_acid_auxiliary_environments": int(auxiliary["environment_id"].nunique()),
    }


def build_astrochemistry(data_dir: Path) -> dict[str, Any]:
    umist = parse_umist_network()
    kida = parse_kida_network()
    network = pd.concat([umist, kida], ignore_index=True).drop_duplicates("reaction_id")
    write_csv(data_dir / "reaction_network.csv", network)
    inventory = build_molecular_inventory(data_dir)
    return {
        "reaction_rows": int(len(network)),
        "networks": network.groupby("source_network").size().astype(int).to_dict(),
        "species_in_reactions": int(len(set(network["reactants"].str.split("+").explode()) | set(network["products"].str.split("+").explode()))),
        "rate_uncertainty_coverage": float(pd.to_numeric(network["uncertainty_factor"], errors="coerce").notna().mean()),
        "molecular_inventory": inventory,
    }


VALID_ELEMENT_SYMBOLS = {
    "H", "He", "Li", "Be", "B", "C", "N", "O", "F", "Ne", "Na", "Mg", "Al", "Si", "P", "S", "Cl", "Ar", "K", "Ca",
    "Sc", "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn", "Ga", "Ge", "As", "Se", "Br", "Kr", "Rb", "Sr",
    "Y", "Zr", "Nb", "Mo", "Tc", "Ru", "Rh", "Pd", "Ag", "Cd", "In", "Sn", "Sb", "Te", "I", "Xe", "Cs", "Ba",
    "La", "Ce", "Pr", "Nd", "Pm", "Sm", "Eu", "Gd", "Tb", "Dy", "Ho", "Er", "Tm", "Yb", "Lu", "Hf", "Ta", "W",
    "Re", "Os", "Ir", "Pt", "Au", "Hg", "Tl", "Pb", "Bi", "Po", "At", "Rn", "Fr", "Ra", "Ac", "Th", "Pa", "U",
}

def element_symbol(isotope: str) -> str | None:
    match = re.match(r"^([A-Z][a-z]?)\-?\d", isotope)
    symbol = match.group(1) if match else None
    return symbol if symbol in VALID_ELEMENT_SYMBOLS else None


def metallicity_from_reference(reference: str) -> float | None:
    match = re.search(r"(\d+(?:\.\d+)?)E-(\d+)", reference)
    if not match:
        return None
    return float(match.group(1)) * 10 ** (-int(match.group(2)))


def build_nucleosynthesis(data_dir: Path) -> dict[str, Any]:
    path = RAW / "SIMPLE_CCSNe_v3p1.hdf5"
    element_rows: list[dict[str, Any]] = []
    isotope_rows: list[dict[str, Any]] = []
    with h5py.File(path, "r") as handle:
        models_group = handle["models"]
        # HDF5 does not guarantee the same iteration order across library versions.
        # Sorting here keeps the canonical CSV byte-identical across Python/h5py runs.
        for model_name in sorted(models_group.keys()):
            group = models_group[model_name]
            decode = lambda name: group[name][0].decode("utf-8") if isinstance(group[name][0], bytes) else str(group[name][0])
            family = decode("dataset")
            citation = decode("citation")
            reference = decode("refid_isoabu")
            mass = float(np.asarray(group["mass"])[0])
            metallicity = metallicity_from_reference(reference)
            mass_coordinate = np.asarray(group["masscoord_mass"], dtype=float)
            shell_mass = np.diff(np.concatenate([[0.0], mass_coordinate]))
            shell_mass = np.clip(shell_mass, 0.0, None)
            abundances = np.asarray(group["abundance_values"], dtype=float)
            isotope_yields = shell_mass @ abundances
            isotopes = [item.decode("utf-8") if isinstance(item, bytes) else str(item) for item in group["abundance_keys"][:]]
            grouped: dict[str, float] = defaultdict(float)
            for isotope, yield_mass in zip(isotopes, isotope_yields):
                element = element_symbol(isotope)
                if element is None or not np.isfinite(yield_mass) or yield_mass <= 0:
                    continue
                grouped[element] += float(yield_mass)
                isotope_rows.append({
                    "source_id": model_name,
                    "model_family": family,
                    "mass_solar": mass,
                    "metallicity": metallicity,
                    "element": element,
                    "isotope": isotope,
                    "yield_mass": float(yield_mass),
                    "uncertainty": np.nan,
                    "initial_abundance_reference": reference,
                    "citation": citation,
                })
            for element, yield_mass in sorted(grouped.items()):
                element_rows.append({
                    "source_id": model_name,
                    "mass_solar": mass,
                    "metallicity": metallicity,
                    "element": element,
                    "yield_mass": yield_mass,
                    "uncertainty": np.nan,
                    "model_family": family,
                    "initial_abundance_reference": reference,
                    "citation": citation,
                })
    element_frame = (
        pd.DataFrame(element_rows)
        .sort_values(["source_id", "element"], kind="mergesort")
        .reset_index(drop=True)
    )
    isotope_frame = (
        pd.DataFrame(isotope_rows)
        .sort_values(["source_id", "element", "isotope"], kind="mergesort")
        .reset_index(drop=True)
    )
    # The HDF5 shell integration can differ around 1e-15 relatively between
    # BLAS thread configurations. Keep full-precision calculations in memory,
    # then serialize a canonical 12-significant-digit scientific value so the
    # published CSV is byte-identical across runners.
    for frame in (element_frame, isotope_frame):
        frame["yield_mass"] = frame["yield_mass"].map(
            lambda value: float(f"{float(value):.12g}")
        )
    write_csv(data_dir / "nucleosynthesis_yields.csv", element_frame)
    write_csv(data_dir / "nucleosynthesis_isotope_yields.csv", isotope_frame)
    return {
        "element_rows": int(len(element_frame)),
        "isotope_rows": int(len(isotope_frame)),
        "models": int(element_frame["source_id"].nunique()),
        "model_families": sorted(element_frame["model_family"].unique().tolist()),
        "masses_solar": sorted(element_frame["mass_solar"].unique().tolist()),
        "metallicity_values": sorted(element_frame["metallicity"].dropna().unique().tolist()),
        "uncertainty": "Non publiée dans ce conteneur; laissée vide.",
    }


def build_isotope_tracers(data_dir: Path) -> dict[str, Any]:
    # Table principale homogène : compilation D/H.
    dh_path = RAW / "Rodriguez_Supplementary_TableS1.xlsx"
    dh = pd.read_excel(dh_path)
    rows: list[dict[str, Any]] = []
    for index, item in dh.iterrows():
        value = numeric(item.get("D/H"))
        if value is None:
            continue
        pos = numeric(item.get("StDev_Pos"))
        neg = numeric(item.get("StDev_Neg"))
        uncertainty = np.nan if pos is None and neg is None else max(pos or 0.0, neg or 0.0)
        sample = str(item.get("Name", "unknown")).strip()
        compound = str(item.get("Compound_Name", "bulk")).strip()
        sample_id = f"{sample}|{compound}|{index + 1}"
        group = "|".join(str(item.get(name, "")).strip() for name in ["Class", "Order", "Subclass", "Type"])
        rows.append({
            "sample_id": sample_id,
            "group": group,
            "tracer": "D/H",
            "value": value,
            "uncertainty": uncertainty,
            "compound_group": item.get("Compound_Group"),
            "compound_class": item.get("Compound_Class"),
            "compound_type": item.get("Compound_Type"),
            "reference": item.get("Reference"),
            "doi": item.get("DOIs"),
        })
    main = pd.DataFrame(rows)
    write_csv(data_dir / "isotope_tracers.csv", main)

    # Mesures lunaires Ca : table auxiliaire, sans mélange dans le clustering D/H.
    ca_path = RAW / "Ca isotope data for analyzed lunar samples_Fu et al 2023.xlsx"
    workbook = load_workbook(ca_path, read_only=True, data_only=True)
    sheet = workbook.active
    ca_rows = []
    category = "unclassified"
    category_labels = {"Lunar rocks", "Low-Ti basalt", "High-Ti basalt", "Highland rock", "Lunar minerals"}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        name = row[0]
        if name in category_labels:
            category = str(name)
            continue
        value = numeric(row[9] if len(row) > 9 else None)
        uncertainty = numeric(row[10] if len(row) > 10 else None)
        if name is not None and value is not None:
            ca_rows.append({"sample_id": str(name), "group": category, "tracer": "delta44_40Ca_permil", "value": value, "uncertainty": uncertainty})
    write_csv(data_dir / "lunar_calcium_isotopes.csv", pd.DataFrame(ca_rows))

    # Ivuna Mn-Cr : résultats synthétiques de chaque feuille, table auxiliaire.
    ivuna_path = RAW / "TableS4.xlsx"
    workbook = load_workbook(ivuna_path, read_only=True, data_only=True)
    cr_rows = []
    for sheet in workbook.worksheets:
        header_row = None
        delta_col = None
        error_col = None
        for ridx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for cidx, value in enumerate(row):
                if str(value).strip() in {"δ53Cr", "d53Cr"}:
                    header_row, delta_col, error_col = ridx, cidx, cidx + 1
                    break
            if header_row:
                break
        if header_row is None:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, max_row=min(header_row + 8, sheet.max_row), values_only=True):
            value = numeric(row[delta_col] if delta_col < len(row) else None)
            uncertainty = numeric(row[error_col] if error_col < len(row) else None)
            if value is not None:
                cr_rows.append({"sample_id": sheet.title, "group": "Ivuna_CI", "tracer": "delta53Cr_permil", "value": value, "uncertainty": uncertainty})
                break
    write_csv(data_dir / "ivuna_mn_cr_isotopes.csv", pd.DataFrame(cr_rows))
    return {
        "dh_rows": int(len(main)),
        "dh_groups": int(main["group"].nunique()),
        "dh_uncertainty_coverage": float(main["uncertainty"].notna().mean()),
        "lunar_ca_rows_auxiliary": len(ca_rows),
        "ivuna_mn_cr_rows_auxiliary": len(cr_rows),
        "scope": "Le clustering canonique utilise seulement la compilation homogène D/H. Ca lunaire et Mn-Cr Ivuna sont conservés séparément.",
    }


def extend_partition_experiments(data_dir: Path) -> dict[str, Any]:
    path = RAW / "Dataset_Fig4_FigS1_FigS2_DC_1.csv"
    source = pd.read_csv(path, encoding="utf-8-sig")
    rows: list[dict[str, Any]] = []
    block_starts = list(range(0, 121, 20))
    for start in block_starts:
        label = str(source.columns[start]).replace("_", " ")
        for row_index, item in source.iterrows():
            pressure = numeric(item.iloc[start + 1])
            temperature = numeric(item.iloc[start + 3])
            delta_iw = numeric(item.iloc[start + 5])
            logd = numeric(item.iloc[start + 9])
            if None in {pressure, temperature, delta_iw, logd}:
                continue
            experiment_name = str(item.iloc[start]).strip() if pd.notna(item.iloc[start]) else ""
            rows.append({
                "experiment_id": experiment_name or f"{label}:row{row_index + 2}",
                "element": "C",
                "pressure_gpa": pressure,
                "temperature_k": temperature,
                "delta_iw": delta_iw,
                "logD": logd,
                "uncertainty": numeric(item.iloc[start + 10]) or np.nan,
                "source_id": label,
                "nbo_t": numeric(item.iloc[start + 7]),
                "h2o": numeric(item.iloc[start + 11]),
                "source_file": path.name,
            })
    existing_path = data_dir / "partition_experiments.csv"
    existing = pd.read_csv(existing_path) if existing_path.exists() else pd.DataFrame()
    extension = pd.DataFrame(rows)
    combined = pd.concat([existing, extension], ignore_index=True, sort=False).drop_duplicates("experiment_id")
    write_csv(existing_path, combined)
    complete = combined[["pressure_gpa", "temperature_k", "delta_iw", "logD"]].notna().all(axis=1)
    return {
        "new_carbon_rows": int(len(extension)),
        "total_rows": int(len(combined)),
        "complete_regression_rows": int(complete.sum()),
        "sources": sorted(combined.get("source_id", pd.Series(dtype=str)).dropna().astype(str).unique().tolist()),
    }


def build_endosymbiosis(data_dir: Path) -> dict[str, Any]:
    zip_path = RAW / "HMM_summarized_results.zip"
    with zipfile.ZipFile(zip_path) as archive:
        text = archive.read("HMM_summarized_results/presence-absence-matrix.csv").decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    next(reader, None)
    records = []
    for row in reader:
        if len(row) < 8:
            continue
        accession, species, section, category, hmm, gene, present, status = row[:8]
        records.append({
            "accession": accession,
            "species": species,
            "section": section,
            "category": category,
            "hmm": hmm,
            "gene": gene,
            "present": 1 if str(present).strip() == "1" else 0,
            "status": status,
        })
    frame = pd.DataFrame(records)
    events = []
    for accession, group in frame.groupby("accession"):
        species = str(group["species"].iloc[0])
        completeness = float(group["present"].mean())
        section_fraction = group.groupby("section")["present"].mean().to_dict()
        metabolic_retention = float(group.loc[group["section"].isin(["PMF", "envelope"]), "present"].mean())
        events.append({
            "event_id": accession,
            "host": "not_reported_in_source",
            "symbiont": species,
            "gene_transfer": np.nan,
            "metabolic_integration": metabolic_retention,
            "dependency": np.nan,
            "evidence_level": np.nan,
            "genome_retention_proxy": completeness,
            "section_retention_json": json.dumps(section_fraction, ensure_ascii=False, sort_keys=True),
            "derived_proxy": True,
        })
    event_frame = pd.DataFrame(events)
    write_csv(data_dir / "endosymbiosis_events.csv", event_frame)
    write_csv(data_dir / "endosymbiont_hmm_presence_absence.csv", frame)

    aphid = pd.read_csv(RAW / "aphid_PGN_count.csv", encoding="utf-8-sig")
    long = aphid.melt(id_vars=[aphid.columns[0]], var_name="host_code", value_name="gene_count")
    long = long.rename(columns={aphid.columns[0]: "gene_family"})
    write_csv(data_dir / "aphid_pgn_gene_counts.csv", long)
    return {
        "genomes": int(len(event_frame)),
        "hmm_rows": int(len(frame)),
        "symbionts": int(event_frame["symbiont"].nunique()),
        "median_retention_proxy": float(event_frame["genome_retention_proxy"].median()),
        "aphid_pgn_rows_auxiliary": int(len(long)),
        "limitations": "Le fichier HMM mesure la rétention de fonctions dans 85 génomes réduits. Les transferts nucléaires et la dépendance directe à l'hôte restent absents.",
    }


def build_auxiliary_tables(data_dir: Path) -> dict[str, Any]:
    summaries: dict[str, Any] = {}
    # Profils de gaz de Murchison : utiles, mais incompatibles avec un bilan fermé noyau-manteau-atmosphère-espace.
    gas_rows = []
    for filename, scenario in [
        ("Murchison-vacc-Zn-GAS.csv", "vacuum"),
        ("Murch-50airZn-GAS.csv", "intermediate_oxidizing"),
        ("Murch-100airZn-GAS.csv", "strong_oxidizing"),
    ]:
        frame = pd.read_csv(RAW / filename, encoding="utf-8-sig")
        for item in frame.itertuples(index=False):
            temperature = numeric(item[0])
            pressure = numeric(item[1])
            for column, value in zip(frame.columns[2:], item[2:]):
                amount = numeric(value)
                if temperature is not None and amount is not None:
                    gas_rows.append({"scenario": scenario, "temperature_k": temperature, "pressure_mpa": pressure, "species": column, "mole_fraction": amount, "source_file": filename})
    gas = pd.DataFrame(gas_rows)
    write_csv(data_dir / "murchison_degassing_profiles.csv", gas)
    summaries["murchison_degassing"] = {"rows": int(len(gas)), "species": int(gas["species"].nunique()), "scenarios": int(gas["scenario"].nunique())}

    meteorite = pd.read_csv(RAW / "Measurement_data.csv", encoding="cp1252", sep=";")
    write_csv(data_dir / "meteorite_thermal_properties.csv", meteorite)
    summaries["meteorite_thermal_properties"] = {"rows": int(len(meteorite)), "meteorites": int(meteorite["Meteorite"].nunique())}

    # Tables noyau/bulk gardées dans leur forme mesurée, sans inventer manteau, atmosphère ou pertes.
    core_bulk = pd.read_excel(RAW / "Dataset_Fig5_bulk_core.xlsx")
    write_csv(data_dir / "core_bulk_h_c_models.csv", core_bulk)
    summaries["core_bulk_h_c_models"] = {"rows": int(len(core_bulk)), "columns": list(core_bulk.columns)}
    return summaries


def source_manifest() -> dict[str, Any]:
    files = []
    for path in sorted(RAW.iterdir()):
        if path.is_file():
            files.append({
                "filename": path.name,
                "bytes": path.stat().st_size,
                "sha256": sha256(path),
                "source_url": SOURCE_URLS.get(path.name),
            })
    return {
        "source_bundle": "aphid_PGN_count.zip fourni par l'utilisateur",
        "integration_date": "2026-08-05",
        "rule": "Chaque fichier conserve son empreinte brute. Aucune licence nouvelle n'est attribuée par ORI-C.",
        "files_retained": files,
    }


def integrate(data_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    if not RAW.exists():
        return {}, {}
    summaries: dict[str, Any] = {}
    summaries["modern_climate_ensemble"] = build_climate_ensemble(data_dir)
    summaries["astrochemistry"] = build_astrochemistry(data_dir)
    summaries["nucleosynthesis_yields"] = build_nucleosynthesis(data_dir)
    summaries["isotope_tracers"] = build_isotope_tracers(data_dir)
    summaries["partition_experiments_extension"] = extend_partition_experiments(data_dir)
    summaries["endosymbiosis_events"] = build_endosymbiosis(data_dir)
    summaries["auxiliary_tables"] = build_auxiliary_tables(data_dir)

    manifest = source_manifest()
    (RAW.parent / "SOURCE.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    rejected = {
        "not_integrated_from_original_bundle": {
            "HadCRUT NetCDF": "Doublon binaire du CSV HadCRUT utilisé.",
            "HMM_raw_results.zip": "La matrice résumée suffit aux métriques retenues; les sorties brutes n'ajoutent pas de variable canonique.",
            "FA_uncertainties_FK.csv": "Le fichier d'abondances correspondant manque dans le lot.",
            "TableS3.xlsx": "Standards et calibration SIMS; conservé en source mais non mélangé aux échantillons naturels.",
            "Dataset_Fig2.xlsx": "Séries sans métadonnées explicites suffisantes dans le fichier pour nommer honnêtement les variables.",
            "210210_MurchisonSteppedHeating_ICPMSData.csv": "Mesures brutes conservées, mais pas converties en bilan volatil fermé sans protocole de normalisation.",
        },
        "canonical_tables_still_absent": ["thermochemical_phases.csv", "late_accretion_tracers.csv", "planetary_histories.csv", "volatile_inventory.csv"],
    }
    (HERE / "TRI_LOT_SCIENTIFIQUE_2026_08_05.json").write_text(json.dumps(rejected, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    coverage = {
        "modern_climate_ensemble": {
            "supported_test_ids": ["CL3-001", "CL3-002", "CL3-003", "CL3-006", "CL3-007", "CL4-001", "CL4-002", "CL4-003", "CL4-005", "CL4-006", "CL4-007"],
            "limitations": "Observations avec incertitude, trajectoires CMIP6 multi-modèles/scénarios et expériences idéalisées. Aucun coût matériel, overshoot explicite, retrait ou restauration n'est fourni.",
        },
        "reaction_network": {
            "supported_test_ids": ["M3-001", "M3-011", "M3-015"],
            "limitations": "Deux réseaux gazeux indépendants avec taux, températures et incertitudes KIDA. Pas de chimie de surface, glaces, ordre d'irradiation ni inventaire radioastronomique.",
        },
        "molecular_inventory": {
            "supported_test_ids": ["M3-001", "M3-011", "M3-015"],
            "limitations": "Conditions initiales Rate22 directement compatibles avec les espèces du réseau. La compilation d'acides aminés est auxiliaire; aucun inventaire radioastronomique n'est prétendu.",
        },
        "nucleosynthesis_yields": {
            "supported_test_ids": ["M2-004"],
            "limitations": "Dix-huit modèles CCSN, six familles et trois masses : utilisables pour l'effet de masse. Pas de BBN, AGB, fusions compactes, rotation/binarité contrôlée ni incertitudes publiées dans le conteneur.",
        },
        "isotope_tracers": {
            "supported_test_ids": ["P1-001"],
            "limitations": "Compilation D/H, mesures Cr d'Ivuna et Ca lunaires avec provenance. Elle compile des traceurs disponibles, mais ne suffit pas à tester la dichotomie carbonée/non carbonée.",
        },
        "partition_experiments": {
            "supported_test_ids": ["P3-001", "P3-002", "P3-003", "P3-004", "P3-005"],
            "limitations": "Compilation étendue par des expériences de partage du carbone avec P, T, redox et logD. Trajectoires planétaires, ordre des apports, océans magmatiques et validation aveugle restent absents.",
        },
        "endosymbiosis_events": {
            "supported_test_ids": ["B2-003"],
            "limitations": "Réduction génomique mesurée par matrice HMM sur 85 génomes. Hôtes, phylogénies, transferts nucléaires, dépendances directes et systèmes d'import protéique ne sont pas reliés dans cette source.",
        },
    }
    for item in coverage.values():
        item["scope_mode"] = "allow_list"
    return summaries, coverage
